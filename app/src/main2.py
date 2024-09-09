import locale
import logging
import os
from selenium.webdriver.edge.service import Service
from selenium import webdriver
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.action_chains import ActionChains
from datetime import datetime, date, timedelta
import time
import smtplib
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import openpyxl
from dotenv import load_dotenv
from trio import sleep

# Cargar variables del .env
load_dotenv()

# Obtener entorno
ENVIRONMENT = os.getenv("ENVIRONMENT", "production")

# Definir rutas según el entorno
CHROMEDRIVER_PATH = os.getenv(f"CHROMEDRIVER_PATH_{ENVIRONMENT.upper()}")
DATA_PATH = os.getenv(f"DATA_PATH_{ENVIRONMENT.upper()}")
EXCEL_PATH = os.getenv(f"EXCEL_PATH_{ENVIRONMENT.upper()}")
INFORMACION_PATH = os.getenv(f"INFORMACION_PATH_{ENVIRONMENT.upper()}")

# Configuración de logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s [%(levelname)s] %(message)s')
logger = logging.getLogger(__name__)

diasDeBusqueda = 3

# Ruta al controlador msedgedriver.exe
service = Service(executable_path=CHROMEDRIVER_PATH)


def resaltar_elemento(driver, elemento):
    """Función para resaltar un elemento cambiando su fondo a amarillo"""
    driver.execute_script("arguments[0].style.backgroundColor = 'yellow'", elemento)


def resaltar_elemento_rojo(driver, elemento):
    driver.execute_script("arguments[0].style.backgroundColor = 'red'", elemento)


def obtenerFechaDeHoy():
    try:
        fechaHoy = datetime.now()
        print(f"Fecha de hoy obtenida: {fechaHoy}", flush=True)
        return fechaHoy
    except Exception as e:
        print("Error obteniendo la fecha de hoy, posible fallo de internet", flush=True)
        return None


def recorrerElExcel(fechaHoy):
    try:
        if not os.path.exists(os.path.dirname(DATA_PATH)):
            print("El directorio /app/data no existe.", flush=True)
            return

        archivo_path = INFORMACION_PATH
        try:
            with open(archivo_path, "w") as archivoActuaciones:
                archivoActuaciones.write("\n")
                archivoActuaciones.flush()
                print(f"Archivo {archivo_path} abierto y limpiado correctamente.", flush=True)
        except IOError as e:
            print(f"No se pudo abrir o escribir en el archivo {archivo_path}: {e}", flush=True)
            return

        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb["CONSULTA UNIFICADA DE PROCESOS"]
        print("Archivo Excel abierto correctamente.", flush=True)
    except Exception as e:
        print(f"Error abriendo el Excel: {e}", flush=True)
        return

    n_filas = ws.max_row
    print(f"Número de filas en la hoja de trabajo: {n_filas}", flush=True)

    try:
        for fila in range(2, n_filas + 1):
            valor = ws['B' + str(fila)].value
            numeroDeProceso = valor
            if valor is None or valor == "":
                continue
            else:
                revisarActuaciones(numeroDeProceso, fechaHoy)
    except Exception as e:
        print(f"Error en recorrer el Excel: {e}", flush=True)

    try:
        with open(archivo_path, "a") as archivoActuaciones:
            archivoActuaciones.write("\n")
            archivoActuaciones.write("######################################\n")
            archivoActuaciones.write(f"# REGISTROS ESCANEADOS: {fila-1} DE: {n_filas-1}\n")
            archivoActuaciones.write("######################################\n")
            archivoActuaciones.flush()
            print(f"Información añadida al archivo {archivo_path}.", flush=True)
    except Exception as e:
        print(f"Error escribiendo en el archivo de información: {e}", flush=True)


def esperarHastaQueNoHayaError(driver):
    while True:
        try:
            error_element = WebDriverWait(driver, 2).until(
                EC.presence_of_element_located((By.XPATH, "//p[text()='Error: Network Error']"))
            )
            if error_element:
                print("Error de red detectado. Reintentando en 10 minutos...", flush=True)
                time.sleep(600)
                driver.refresh()
            else:
                print("No hay error de red. Continuando con el flujo.", flush=True)
                break
        except TimeoutException:
            print("No se encontró el texto 'Error: Network Error'. Continuando con el flujo.",
                  flush=True)
            break


def revisarActuaciones(numeroDeProceso, fechaHoy):

    options = webdriver.ChromeOptions()
    options.add_argument('--headless')  # Activar el modo headless
    options.add_argument('--disable-gpu')  # Necesario para algunos sistemas
    options.add_argument('--no-sandbox')  # Previene errores en ciertos entornos
    options.add_argument('--disable-dev-shm-usage')  # Mejora la estabilidad en entornos con poca memoria compartida

    driver = webdriver.Chrome(service=service, options=options)

    while True:
        try:
            time.sleep(1)
            driver.get("https://consultaprocesos.ramajudicial.gov.co/Procesos/NumeroRadicacion")
            time.sleep(1)
            print("Página cargada exitosamente.", flush=True)
            break
        except Exception as e:
            print(f"Error al cargar la página: {e}. Reintentando...", flush=True)
            driver.refresh()
            time.sleep(300)
            if not driver.window_handles:
                print("Se detuvo el escaneo de registros manualmente.", flush=True)
                break

    try:
        div_elements = driver.find_elements(By.CSS_SELECTOR, "div.v-input--selection-controls__input")
        second_div_element = div_elements[2]
        resaltar_elemento(driver, second_div_element)
        second_div_element.click()

        element = WebDriverWait(driver, 60).until(
            EC.presence_of_element_located(
                (By.XPATH, "//input[@placeholder='Ingrese los 23 dígitos del número de Radicación']")
            )
        )
        resaltar_elemento(driver, element)
        time.sleep(1)
        element.send_keys(numeroDeProceso)
        print(f"Número de radicación ingresado: {element.get_attribute('value')}", flush=True)

        span_consultar = WebDriverWait(driver, 60).until(
            EC.presence_of_element_located((By.XPATH, "//span[text()='Consultar']"))
        )
        resaltar_elemento(driver, span_consultar)
        time.sleep(1)
        span_consultar.click()

        try:
            boton_volver = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, "//span[text()=' Volver ']"))
            )
            if boton_volver:
                esperarHastaQueNoHayaError(driver)
                resaltar_elemento(driver, boton_volver)
                boton_volver.click()
                print("Botón 'Volver' encontrado y clickeado.", flush=True)

        except TimeoutException:
            print("No hay span 'Volver'.", flush=True)
            pass

        try:
            tablas = WebDriverWait(driver, 40).until(
                EC.presence_of_all_elements_located((By.TAG_NAME, "table"))
            )

            resTablaFechaAC = tablas[0]
            filas = resTablaFechaAC.find_elements(By.TAG_NAME, "tr")

            for fila in filas:
                celdas = fila.find_elements(By.TAG_NAME, "td")
                if len(celdas) > 0:
                    resaltar_elemento_rojo(driver, fila)

                    try:
                        boton = celdas[2].find_element(By.TAG_NAME, "button")
                        resaltar_elemento(driver, boton)

                        fecha_boton_texto = boton.text.strip()
                        fecha_boton = date.fromisoformat(fecha_boton_texto)

                        for i in range(0, diasDeBusqueda):
                            fecha_comparacion = fechaHoy - timedelta(days=i)
                            fecha_comparacion_str = fecha_comparacion.strftime("%Y-%m-%d")
                            if fecha_boton_texto == fecha_comparacion_str:
                                print(f"Fecha {fecha_boton_texto} encontrada en el rango.", flush=True)
                                boton.click()
                                break
                        else:
                            print(f"Fecha {fecha_boton_texto} no está en el rango de búsqueda.", flush=True)

                    except Exception as e:
                        print(f"No se encontró el botón en esta fila", flush=True)

        except Exception as e:
            print(f"Error en el recorrido de la tabla o búsqueda del botón: {e}", flush=True)

    except Exception as e:
        print(f"Error en la búsqueda de tabla: {e}", flush=True)
        pass

    try:

        tablas = WebDriverWait(driver, 60).until(

            EC.presence_of_all_elements_located((By.TAG_NAME, "table"))

        )

        time.sleep(1)

        tabla_actuaciones = tablas[1]

        for fila in tabla_actuaciones.find_elements(By.TAG_NAME, "tr"):

            if len(fila.find_elements(By.TAG_NAME, "td")) > 0:

                celdaFecha = fila.find_elements(By.TAG_NAME, "td")[0]

                fechaObtenida = date.fromisoformat(celdaFecha.text)

                actuacionObtenida = fila.find_elements(By.TAG_NAME, "td")[1].text

                anotacionObtenida = fila.find_elements(By.TAG_NAME, "td")[2].text

                fechaObtenida_str = fechaObtenida.strftime("%Y-%m-%d")

                for i in range(0, diasDeBusqueda):

                    fecha_comparacion = fechaHoy - timedelta(days=i)

                    fecha_comparacion = fecha_comparacion.strftime("%Y-%m-%d")

                    if fechaObtenida_str == fecha_comparacion:
                        with open(INFORMACION_PATH, "a") as archivoActuaciones:
                            archivoActuaciones.write("\n")

                            archivoActuaciones.write(f"NUMERO DEL PROCESO: {numeroDeProceso}\n")

                            archivoActuaciones.write(f"Fecha: {fechaObtenida_str}\n")

                            archivoActuaciones.write(f"Actuacion: {actuacionObtenida}\n")

                            archivoActuaciones.write(f"Anotacion: {anotacionObtenida}\n")

                            archivoActuaciones.write(

                                "-----------------------------------------------------------------\n")

                        print(f"Información guardada para la fecha: {fechaObtenida_str}", flush=True)

    except Exception as e:

        print(f"No se realizó la búsqueda de actuaciones", flush=True)

        pass

    finally:

        driver.quit()


def enviarArchivoCorreo():
    fecha_ahora = datetime.now()
    fechaHoyStr = fecha_ahora.strftime("%A %d - %m - %Y a las %I:%M %p").capitalize()

    smtp = smtplib.SMTP_SSL("smtp.gmail.com")
    correo_emisor = "registroautomaticoactuaciones@gmail.com"
    correo_receptor = "registroautomaticoactuaciones@gmail.com"
    asunto = "Registro escaneo de documento desde el servidor"

    archivo_adjuntar = INFORMACION_PATH

    smtp.login(correo_emisor, "lctc zggr fztd eokc")

    mensaje = MIMEMultipart()
    mensaje["Subject"] = asunto
    mensaje["From"] = correo_emisor
    mensaje["To"] = correo_receptor

    cuerpo = f"Escaneo de actuaciones\n"
    mensaje.attach(MIMEText(cuerpo))

    with open(archivo_adjuntar, "rb") as archivo:
        archivo_adjunto = MIMEApplication(archivo.read())
        archivo_adjunto.add_header('Content-Disposition', 'attachment', filename=archivo_adjuntar)
        mensaje.attach(archivo_adjunto)

    smtp.sendmail(correo_emisor, correo_receptor, mensaje.as_string())
    smtp.quit()
    print("Correo enviado exitosamente.", flush=True)


def main():
    try:
        fechaHoy = obtenerFechaDeHoy()

        recorrerElExcel(fechaHoy)

        enviarArchivoCorreo()

    except Exception as e:
        print(f"Error en el main: {e}", flush=True)


if __name__ == "__main__":
    main()
