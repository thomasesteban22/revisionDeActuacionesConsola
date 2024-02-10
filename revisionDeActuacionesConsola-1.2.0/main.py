from selenium import webdriver
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

import datetime
import locale
from datetime import date, timedelta
from datetime import datetime
from time import sleep

import smtplib
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

import logging
import openpyxl

import tkinter as tk
import tkinter.filedialog as filedialog


import requests

def check_internet():
    try:
        requests.get("https://consultaprocesos.ramajudicial.gov.co/Procesos/NumeroRadicacion", timeout=10)  # Agregué un tiempo de espera de 5 segundos
        return True
    except (requests.ConnectionError, requests.Timeout) as exception:
        return False


diasDeBusqueda = int(input("¿Cuántos días quiere escanear?\n"))

def reporteDeErrores(mensaje):
    fechaHoy = obtenerFechaDeHoy
    fechaHoyStr = str(fechaHoy)
    archivoReporteErrores = open("reporteDeErrores.txt", "a")
    archivoReporteErrores.write("\n") 
    archivoReporteErrores.write("FECHA: ", fechaHoyStr, " ERROR: ", mensaje)
    archivoReporteErrores.write("\n") 


def guardadoDeLogs(fInicio, fFinal, numRegistrosEscaneados):
    logging.info("Fecha Inicio: ", fInicio.strftime("%d/%m/%Y %H:%M:%S"))
    logging.info("Fecha Final: ", fFinal.strftime("%d/%m/%Y %H:%M:%S"))
    logging.info("Registros Escaneados: ", numRegistrosEscaneados)

def obtenerFechaDeHoy():
    try:
        fechaHoy = datetime.now()
        return fechaHoy
    except:
        mensaje = "Hubo un fallo en el internet"
        reporteDeErrores(mensaje)
        print("Error obteniendo la fecha de hoy, posible fallo de internet")

def recorrerElExcel():
    try:
        # Abre el archivo Excel
        wb = openpyxl.load_workbook("FOLDERESBASENUEVA.xlsm")
        # Obtiene la hoja de trabajo activa
        ws = wb.get_sheet_by_name("CONSULTA UNIFICADA DE PROCESOS")
    except:
        mensaje = "Hubo un fallo en el Excel, revisar Excel"
        reporteDeErrores(mensaje)
        print("Error abriendo el excel, verificar ruta y nombre del archivo")

    # Obtiene el número de filas de la hoja de trabajo
    n_filas = ws.max_row
    fila = 2
    # Inicio de lectura en bucle
    try:
        while fila < n_filas:
            for fila in range(2, n_filas + 1):
                # Obtiene el valor de la celda A de la fila actual
                valor = ws['B' + str(fila)].value
                numeroDeProceso = valor     
                # Si el valor es None, pasa
                if valor is None or valor == "":
                        pass
                else:
                    revisarActuaciones(numeroDeProceso)
            fila = fila + 1
            print("Numero de proceso: " + str(fila) + " de "+ str(n_filas))
    except:
        pass

    archivoActuaciones = open("informacion.txt", "a")
    archivoActuaciones.write("\n")

    indiceFila = fila - 1

    texto ="# " +"REGISTROS ESCANEADOS: " + str(indiceFila) + " DE: " + str(n_filas) + "\n"
   
    # Escribir el texto al archivo
    archivoActuaciones.write("#####################################" + "\n")
    archivoActuaciones.write(texto)
    archivoActuaciones.write("#####################################" + "\n")
    archivoActuaciones.close()



def revisarActuaciones(numeroDeProceso):
    print(numeroDeProceso)
    #Obtiene la fecha de hoy
    fechaHoy = obtenerFechaDeHoy()
    # Crea un objeto `EdgeOptions`
    options = webdriver.EdgeOptions()

    # Abre el navegador
    driver = webdriver.Edge(options=options)


    
    try:
        while True:
            if check_internet():
                try:
                    driver.get("https://consultaprocesos.ramajudicial.gov.co/Procesos/NumeroRadicacion")
                    driver.maximize_window()

                    # Espera hasta que se cargue la página
                    while driver.title != "Consulta de Procesos por Número de Radicación- Consejo Superior de la Judicatura":
                        sleep(5)
                        driver.get("https://consultaprocesos.ramajudicial.gov.co/Procesos/NumeroRadicacion")
                        driver.maximize_window()

                    break  # Sal del bucle si la página se cargó correctamente

                except Exception as e:
                    print(f"Error al cargar la página: {e}")

            else:
                print("No hay conexión a internet. Reintentamos en 30 segundos...")
                sleep(30)
    except:
        mensaje = "No se pudo acceder a la pagina de la rama, fallo de internet o mantenimiento"
        reporteDeErrores(mensaje)
        driver.refresh()
        if driver.window_handles == []:
            print("Se Se detuvo el escaneo de registros manualmente")
        print("Error al cargar la pagina, posible fallo de internet")

    
            

    try:
        div_elements = driver.find_elements(By.CSS_SELECTOR, "div.v-input--selection-controls__input")
        second_div_element = div_elements[2]
        second_div_element.click()
    except:
        mensaje = "No se pudo obtener el select Todos los procesos"
        #reporteDeErrores(mensaje)
        print("Error obteniendo la seleccion de todos los procesos")


    try:    
        element = WebDriverWait(driver, 60).until(
            EC.presence_of_element_located((By.XPATH, "//input[@placeholder='Ingrese los 23 dígitos del número de Radicación']"))
        )
        # Escribe el número de radicación en el elemento
        element.send_keys(numeroDeProceso)
        # Obtiene el texto del input
        numero_radicacion = element.text
        print(numero_radicacion)
    except:
        mensaje = "No se pudo obtener el input de 23 digitos"
        #reporteDeErrores(mensaje)
        print("error en el input de los 23 digitos")


    try:
        # Hace click en el botón "Consultar"
        span_consultar = WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.XPATH, "//span[text()='Consultar']")))
        sleep(1)
        # Haz click en el span "Consultar"
        span_consultar.click()
    except:
        mensaje = "No se pudo obtener el span consultar"
        #reporteDeErrores(mensaje)
        print("Error en el boton consultar")

    try:
        boton_volver = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, "//span[text()=' Volver ']"))
        )

        if boton_volver is not None:
            driver.execute_script("arguments[0].style.backgroundColor = 'red';", boton_volver)
        # Haz clic en el botón
            boton_volver.click()
    except:
       print("No hay span volver")
       pass

    try:
        tablas = WebDriverWait(driver, 40).until(
        EC.presence_of_all_elements_located((By.TAG_NAME, "table"))
        )
        resTablaFechaAC = tablas[0]
        for fila in resTablaFechaAC.find_elements(By.TAG_NAME, "tr"):
            driver.execute_script("arguments[0].style.backgroundColor = 'yellow';", fila)
            if len(fila.find_elements(By.TAG_NAME, "td")) > 0:
                celdaFecha = fila.find_elements(By.TAG_NAME, "td")[2]
                driver.execute_script("arguments[0].style.backgroundColor = 'red';", celdaFecha)
            
                botonFecha = celdaFecha.find_element(By.TAG_NAME, "button")
                if(botonFecha is not None):
                    driver.execute_script("arguments[0].style.backgroundColor = 'green';", botonFecha)
                    print(botonFecha.text)
                    fechaInicialComparar = date.fromisoformat(botonFecha.text)
                    fechaInicialStr = fechaInicialComparar.strftime("%Y-%m-%d")
                    for i in range(0, diasDeBusqueda):
                        fechaTemporalComprar = fechaHoy - timedelta(days=i)
                        fechaTemporalComprar = fechaTemporalComprar.date()
                        fechaTemporalComprar = fechaTemporalComprar.strftime("%Y-%m-%d")
                        if(fechaInicialStr == fechaTemporalComprar):
                            botonFecha.click()
                            break
                        else:
                            pass
                            #print("No hay fechas que coincidan")
                    
    except:
        mensaje = "No se pudo obtener la tabla con las ultimas actuaciones"
        #reporteDeErrores(mensaje)
        print("Error en la busqueda de tabla")
        pass

    
    try:
        # Espera a que las tablas se carguen
        tablas = WebDriverWait(driver, 60).until(
        EC.presence_of_all_elements_located((By.TAG_NAME, "table"))
        )
        # Encuentra la tabla "ACTUACIONES"
        tabla_actuaciones = tablas[1]

        # Itera sobre las filas de la tabla
        for fila in tabla_actuaciones.find_elements(By.TAG_NAME, "tr"):

            # Verifica que la lista no esté vacía
            if len(fila.find_elements(By.TAG_NAME, "td")) > 0:

                    # Obtiene el primer td de la fila
                celdaFecha = fila.find_elements(By.TAG_NAME, "td")[0]
                celdaActuacion = fila.find_elements(By.TAG_NAME, "td")[1]
                celdaAnotacion = fila.find_elements(By.TAG_NAME, "td")[2]

                    # Subraya la celdaFecha en amarillo
                driver.execute_script("arguments[0].style.backgroundColor = 'yellow';", celdaFecha)
                fechaObtenida = date.fromisoformat(celdaFecha.text)
                actuacionObtenida = celdaActuacion.text
                anotacionObtenida = celdaAnotacion.text
                fechaObtenida_str = fechaObtenida.strftime("%Y-%m-%d")
                print(fechaObtenida_str, "OBTENIDA")
                    
                for i in range(0, diasDeBusqueda):
                    fecha_comparacion = fechaHoy - timedelta(days=i)
                    fecha_comparacion = fecha_comparacion.date()
                    fecha_comparacion = fecha_comparacion.strftime("%Y-%m-%d")
                    if fechaObtenida_str == fecha_comparacion:
                            archivoActuaciones = open("informacion.txt", "a")
                            print("FECHA COMPATIBLE ------------------------->")
                            archivoActuaciones.write("\n")                                                                                                                                                  
                            archivoActuaciones.write("NUMERO DEL PROCESO: " + numeroDeProceso + "\n")
                            archivoActuaciones.write("Fecha: " + fechaObtenida_str + "\n")
                            archivoActuaciones.write("Actuacion: " + actuacionObtenida + "\n")
                            archivoActuaciones.write("Anotacion: " + anotacionObtenida + "\n")
                            archivoActuaciones.write("---------------------------------------------------------------------------------------------------------------------------"+ "\n")
                    else:
                        pass
        archivoActuaciones.close()            
        driver.quit()
    except:
        print("No se realizo la busqueda de actuaciones")
        pass
        

def enviarArchivoCorreo():
    fecha_ahora = datetime.now()
    locale.setlocale(locale.LC_ALL, 'es_ES')
    fechaHoyStr = fecha_ahora.strftime("%A %d - %m - %Y a las %I:%M %p")
    fechaHoyStr = fechaHoyStr.capitalize()

    smtp = smtplib.SMTP_SSL("smtp.gmail.com")

    correo_emisor = "registroautomaticoactuaciones@gmail.com"
    correo_receptor = "registroautomaticoactuaciones@gmail.com"
    asunto = "Registro escaneo de documento"

    archivo_adjuntar = "informacion.txt"

    smtp.login(correo_emisor, "lctc zggr fztd eokc")
 
    mensaje = MIMEMultipart()
    mensaje["Subject"] = asunto
    mensaje["From"] = correo_emisor
    mensaje["To"] = correo_receptor

    cuerpo = "Correo Auto-generado, " + fechaHoyStr
    mensaje.attach(MIMEText(cuerpo, "plain"))

    archivo = open(archivo_adjuntar, "rb")
    adjunto = MIMEApplication(archivo.read())
    adjunto.add_header("Content-Disposition", "attachment; filename={}".format(archivo_adjuntar))
    mensaje.attach(adjunto)

    smtp.sendmail(correo_emisor, correo_receptor, mensaje.as_string())
    archivoActuaciones = open("informacion.txt", "w")
    archivoActuaciones.write("")
    archivoActuaciones.close()
    print("Correo enviado...")
    smtp.quit()
    
def main():
  
    logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
    filename="registros.log",
    )
    try:
        recorrerElExcel()
    except Exception as e:
        pass
        print(e.args[0])
        print("Se detuvo el recorrido de Excel")

    try:
        enviarArchivoCorreo()
    except Exception as e:
        mensaje = "No se envio el correo, fallo de internet"
        #reporteDeErrores(mensaje)
        print(e.args[0])
        print("Error en el envio del correo")

main()