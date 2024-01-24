from selenium import webdriver
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

import datetime
import locale
from datetime import date, timedelta
from datetime import datetime
from time import sleep

import smtplib
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

import logging
import openpyxl

import tkinter as tk
import tkinter.filedialog as filedialog
from tkinter import ttk

import termcolor


def guardadoDeLogs(fInicio, fFinal, numRegistrosEscaneados):
    logging.info("Fecha Inicio: ", fInicio.strftime("%d/%m/%Y %H:%M:%S"))
    logging.info("Fecha Final: ", fFinal.strftime("%d/%m/%Y %H:%M:%S"))
    logging.info("Registros Escaneados: ", numRegistrosEscaneados)

def obtenerFechaDeHoy():
    try:
        fechaHoy = datetime.now()
        return fechaHoy
    except:
        print("Error obteniendo la fecha de hoy, posible fallo de internet")

def recorrerElExcel():
    try:
        # Abre el archivo Excel
        wb = openpyxl.load_workbook("testBaseDeDatos.xlsx")
        # Obtiene la hoja de trabajo activa
        ws = wb.active
    except:
        print("Error abriendo el excel, verificar ruta y nombre del archivo")

    # Obtiene el número de filas de la hoja de trabajo
    n_filas = ws.max_row
    # Inicio de lectura en bucle
    try:
        for fila in range(2, n_filas + 1):
            # Obtiene el valor de la celda A de la fila actual
            valor = ws['D' + str(fila)].value
            numeroDeProceso = valor     
            # Si el valor es None, pasa
            if valor is None or valor == "":
                    pass
            else:
                revisarActuaciones(numeroDeProceso)
    except:
        pass

    archivoActuaciones = open("informacion.txt", "a")
    archivoActuaciones.write("\n")

    indiceFila = fila - 1

    texto ="# " +"REGISTROS ESCANEADOS: " + str(indiceFila) + " DE: " + str(n_filas) + "\n"
   
    # Escribir el texto al archivo
    archivoActuaciones.write("#####################################" + "\n")
    archivoActuaciones.write(texto)
    archivoActuaciones.write("#####################################" + "\n")
    archivoActuaciones.close()



def revisarActuaciones(numeroDeProceso):
    print(numeroDeProceso)
    #Obtiene la fecha de hoy
    fechaHoy = obtenerFechaDeHoy()
    # Crea un objeto `EdgeOptions`
    options = webdriver.EdgeOptions()

    # Abre el navegador
    driver = webdriver.Edge(options=options)

    try:
        # Abre la página web
        driver.get("https://consultaprocesos.ramajudicial.gov.co/Procesos/NumeroRadicacion")
        driver.maximize_window()
        sleep(1)
    except:
        if driver.window_handles == []:
            print("Se Se detuvo el escaneo de registros manualmente")
        print("Error al cargar la pagina, posible fallo de internet")

        # Encuentra el elemento que deseas hacer clic
    div_elements = driver.find_elements(By.CSS_SELECTOR, "div.v-input--selection-controls__input")
    second_div_element = div_elements[2]
    second_div_element.click()


    element = WebDriverWait(driver, 60).until(
        EC.presence_of_element_located((By.XPATH, "//input[@placeholder='Ingrese los 23 dígitos del número de Radicación']"))
    )
    sleep(1)
    # Escribe el número de radicación en el elemento
    element.send_keys(numeroDeProceso)
    # Obtiene el texto del input
    numero_radicacion = element.text
    print(numero_radicacion)

    # Hace click en el botón "Consultar"
    span_consultar = WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.XPATH, "//span[text()='Consultar']")))
    sleep(1)
    # Haz click en el span "Consultar"
    span_consultar.click()

    try:
        boton_volver = WebDriverWait(driver, 30).until(
        EC.presence_of_element_located((By.XPATH, "//span[text()=' Volver ']"))
        )

        if boton_volver is not None:
            driver.execute_script("arguments[0].style.backgroundColor = 'red';", boton_volver)
        # Haz clic en el botón
            boton_volver.click()
    except:
       print("No hay span volver")
       pass

    
    try:
        tablas = WebDriverWait(driver, 60).until(
        EC.presence_of_all_elements_located((By.TAG_NAME, "table"))
        )
        resTablaFechaAC = tablas[0]
        for fila in resTablaFechaAC.find_elements(By.TAG_NAME, "tr"):
            driver.execute_script("arguments[0].style.backgroundColor = 'yellow';", fila)
            sleep(1)
            if len(fila.find_elements(By.TAG_NAME, "td")) > 0:
                celdaFecha = fila.find_elements(By.TAG_NAME, "td")[2]
                sleep(1)
                driver.execute_script("arguments[0].style.backgroundColor = 'red';", celdaFecha)
            
                botonFecha = celdaFecha.find_element(By.TAG_NAME, "button")
                if(botonFecha is not None):
                    driver.execute_script("arguments[0].style.backgroundColor = 'green';", botonFecha)
                    print(botonFecha.text)
                    fechaInicialComparar = date.fromisoformat(botonFecha.text)
                    fechaInicialStr = fechaInicialComparar.strftime("%Y-%m-%d")
                    for i in range(0, 100):
                        fechaTemporalComprar = fechaHoy - timedelta(days=i)
                        fechaTemporalComprar = fechaTemporalComprar.date()
                        fechaTemporalComprar = fechaTemporalComprar.strftime("%Y-%m-%d")
                        if(fechaInicialStr == fechaTemporalComprar):
                            botonFecha.click()
                            break
                        else:
                            pass
                            #print("No hay fechas que coincidan")
                    
    except:
        print("Error en la busqueda de tabla")
        pass

    
    try:
        # Espera a que las tablas se carguen
        tablas = WebDriverWait(driver, 60).until(
        EC.presence_of_all_elements_located((By.TAG_NAME, "table"))
        )
        sleep(1)
        # Encuentra la tabla "ACTUACIONES"
        tabla_actuaciones = tablas[1]

        # Itera sobre las filas de la tabla
        for fila in tabla_actuaciones.find_elements(By.TAG_NAME, "tr"):

            # Verifica que la lista no esté vacía
            if len(fila.find_elements(By.TAG_NAME, "td")) > 0:

                    # Obtiene el primer td de la fila
                celdaFecha = fila.find_elements(By.TAG_NAME, "td")[0]
                celdaActuacion = fila.find_elements(By.TAG_NAME, "td")[1]
                celdaAnotacion = fila.find_elements(By.TAG_NAME, "td")[2]

                    # Subraya la celdaFecha en amarillo
                driver.execute_script("arguments[0].style.backgroundColor = 'yellow';", celdaFecha)
                fechaObtenida = date.fromisoformat(celdaFecha.text)
                actuacionObtenida = celdaActuacion.text
                anotacionObtenida = celdaAnotacion.text
                fechaObtenida_str = fechaObtenida.strftime("%Y-%m-%d")
                print(fechaObtenida_str, "OBTENIDA")
                    
                for i in range(0, 100):
                    fecha_comparacion = fechaHoy - timedelta(days=i)
                    fecha_comparacion = fecha_comparacion.date()
                    fecha_comparacion = fecha_comparacion.strftime("%Y-%m-%d")
                    if fechaObtenida_str == fecha_comparacion:
                            archivoActuaciones = open("informacion.txt", "a")
                            print("FECHA COMPATIBLE ------------------------->")
                            archivoActuaciones.write("\n")                                                                                                                                                  
                            archivoActuaciones.write("NUMERO DEL PROCESO: " + numeroDeProceso + "\n")
                            archivoActuaciones.write("Fecha: " + fechaObtenida_str + "\n")
                            archivoActuaciones.write("Actuacion: " + actuacionObtenida + "\n")
                            archivoActuaciones.write("Anotacion: " + anotacionObtenida + "\n")
                            archivoActuaciones.write("-----------------------------------------------------------------"+ "\n")
                    else:
                        pass
        archivoActuaciones.close()            
        driver.quit()
    except:
        print("No se realizo la busqueda de actuaciones")
        pass
        

def enviarArchivoCorreo():
    fecha_ahora = datetime.now()
    locale.setlocale(locale.LC_ALL, 'es_ES')
    fechaHoyStr = fecha_ahora.strftime("%A %d - %m - %Y a las %I:%M %p")
    fechaHoyStr = fechaHoyStr.capitalize()

    smtp = smtplib.SMTP_SSL("smtp.gmail.com")

    correo_emisor = "registroautomaticoactuaciones@gmail.com"
    correo_receptor = "registroautomaticoactuaciones@gmail.com"
    asunto = "Registro escaneo de documento"

    archivo_adjuntar = "informacion.txt"

    smtp.login(correo_emisor, "lctc zggr fztd eokc")
 
    mensaje = MIMEMultipart()
    mensaje["Subject"] = asunto
    mensaje["From"] = correo_emisor
    mensaje["To"] = correo_receptor

    cuerpo = "Correo Auto-generado, " + fechaHoyStr
    mensaje.attach(MIMEText(cuerpo, "plain"))

    archivo = open(archivo_adjuntar, "rb")
    adjunto = MIMEApplication(archivo.read())
    adjunto.add_header("Content-Disposition", "attachment; filename={}".format(archivo_adjuntar))
    mensaje.attach(adjunto)

    smtp.sendmail(correo_emisor, correo_receptor, mensaje.as_string())
    archivoActuaciones = open("informacion.txt", "w")
    archivoActuaciones.write("")
    archivoActuaciones.close()
    print("Correo enviado...")
    smtp.quit()
    
def main():
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s",
        filename="registros.log",
    )
 
    try:
        recorrerElExcel()
    except Exception as e:
        pass
        print(e.args[0])
        print("Se detuvo el recorrido de Excel")
   
    try:
        enviarArchivoCorreo()
    except Exception as e:
        print(e.args[0])
        print("Error en el envio del correo")


#Ejecuccion del programa
main()